#!/usr/bin/env python3
"""Ring-light Variable Lighting: torch flash/no-flash illuminant + CAT arms.

Compares Bradford CAT white-point arms on the Pansor torch zip cohort:
  - frozen_5500    — claimable baseline (current pipeline)
  - lu_cct         — Lu ambient CCT + MK350 torch CCT prior (no ECC / SPD RGB)
  - lu_spd_ecc     — Lu + measured torch SPD RGB + ECC flash alignment
  - lu_fused       — 80% Lu + 20% no-flash cheek chroma CCT
  - hybrid_deploy  — Lu on F12/warm; frozen 5500 K on D65-like
  - mk350_ring_xy  — oracle in-situ MK350 ring chromaticity at face

FitSkin forehead Lab (per person × ring illuminant) from Booth Lighting.xlsx.
Cheek pipeline: pre-AWB reflectance → tier3 affine → CAT → FairFace7 ROI sampling.

Example:
  python3 scripts/evaluate_ringlight_torch_illuminant.py \\
    --data-root ~/Downloads/Variable\\ Lighting\\ Ring\\ Light-.../Variable\\ Lighting\\ Ring\\ Light \\
    --booth-xlsx ~/Downloads/Booth\\ Lighting.xlsx \\
    --torch-dir ~/Downloads/Torch_meas-.../Torch_meas \\
    --out-dir results/torch_illuminant_ringlight
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from delta_e_2000 import delta_e_2000  # noqa: E402
from flash_noflash_spectral import planck_xyz_y1  # noqa: E402
from models.fairface_race import FairFacePredictor, face_rgb_crop_from_landmarks  # noqa: E402
from pipeline.skin_roi import apple_face_skin_roi_mask  # noqa: E402
from scripts.evaluate_pansor20_chartfree_d65 import (  # noqa: E402
    FOREHEAD_SKIN_LAB_TRIM,
    extract_zip,
    linear_rgb_to_preview_bgr,
    load_affine,
    load_apple_landmarks,
    load_dng_linear,
    match_flash_exposure,
    mean_lab_on_mask,
)
from pipeline.post_corrections import (  # noqa: E402
    apply_lab_affine,
    apply_multi_illuminant_lab_affine,
    load_color_projector,
    load_lab_affine,
    load_multi_illuminant_lab_affine,
    select_affine_for_illuminant,
)
from pipeline.illuminant_estimation import (  # noqa: E402
    ORACLE_CCT,
    TorchPrior,
    align_flash_linear,
    chroma_to_planck_cct,
    estimate_lu,
    fused_lu_nf_cct,
    hybrid_deploy_cct,
    illuminant_vs_mk350,
    load_torch_prior,
    parse_spectrum_file,
    scene_white_for_cat,
    xy_to_cct_mccamy,
    xy_white_y1,
)

PERSON_ALIASES = {
    "lihn": "Lihn",
    "linh": "Lihn",
    "parker": "Parker",
    "anjana": "Anjana",
    "ariana": "Anjana",
    "arjana": "Anjana",
    "woojae": "Woojae",
    "wooj": "Woojae",
}

ORACLE_CCT = ORACLE_CCT  # re-export for callers
DEFAULT_RING_XY = {
    "D65": (0.309932, 0.324021),
    "F12": (0.4297, 0.3914),
}

# Primary CAT arms for improved illuminant estimation experiments
PRIMARY_ARMS = [
    "frozen_5500",
    "lu_cct",
    "lu_spd_ecc",
    "lu_fused",
    "hybrid_deploy",
    "hybrid_lab_affine",
    "hybrid_color_projector",
    "hybrid_multi_lab",
    "hybrid_multi_lab_allfit",
    "hybrid_routed_affine",
    "hybrid_routed_affine_multi_lab",
    "mk350_ring_xy",
]


def in_frame_cc_white_scale(r0: np.ndarray, preview_bgr: np.ndarray) -> Optional[float]:
    """Per-trial exposure from in-frame MCC white patch (torch zips include CC)."""
    import physio_skin_lab_raw_pr250 as pr250
    from mcc24_canonical_d65 import WHITE_PATCH_INDEX, load_canonical_xyz_d65

    got = pr250.patch_linear_rgb_24(r0, preview_bgr, use_median=True)
    if got is None:
        return None
    patches, _ = got
    pw = patches[WHITE_PATCH_INDEX]
    y_cam = 0.2126 * pw[0] + 0.7152 * pw[1] + 0.0722 * pw[2]
    xyz_white = load_canonical_xyz_d65()[WHITE_PATCH_INDEX] / 100.0
    return float(xyz_white[1] / max(y_cam, 1e-12))


def load_torch_spd_prior(torch_dir: Path) -> Dict[str, Any]:
    """Backward-compatible dict wrapper around :class:`TorchPrior`."""
    tp = load_torch_prior(torch_dir)
    return {
        "torch_cct_k": tp.torch_cct_k,
        "flash_rgb": tp.flash_rgb.tolist(),
        "n_files": tp.n_files,
        "files": list(tp.files),
    }


def resolve_downloads_path(
    pattern: str,
    *,
    must_contain: Optional[str] = None,
    downloads: Path = Path.home() / "Downloads",
) -> Optional[Path]:
    """Find the newest Downloads folder matching a glob (no ``...`` placeholders)."""
    if not downloads.is_dir():
        return None
    hits = sorted(downloads.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    for hit in hits:
        if must_contain and must_contain not in str(hit):
            continue
        return hit
    return None


def default_data_root() -> Path:
    hit = resolve_downloads_path("Variable Lighting Ring Light*/Variable Lighting Ring Light")
    if hit:
        return hit
    return Path(
        "/home/mabl-main/Downloads/Variable Lighting Ring Light-20260829T185351Z-1-001/Variable Lighting Ring Light"
    )


def load_ring_illuminant_xy(
    *,
    booth_xlsx: Optional[Path] = None,
    lighting_info_dir: Optional[Path] = None,
) -> Dict[str, Tuple[float, float]]:
    """MK350-measured ring chromaticity per illuminant label (D65 / F12)."""
    out = dict(DEFAULT_RING_XY)
    if booth_xlsx and Path(booth_xlsx).is_file():
        from openpyxl import load_workbook

        wb = load_workbook(booth_xlsx, data_only=True)
        for sn in wb.sheetnames:
            if sn.strip().lower() not in ("linh", "lihn"):
                continue
            for row in wb[sn].iter_rows(values_only=True):
                if row and len(row) >= 9 and row[6] in ("D65", "F12"):
                    try:
                        out[str(row[6])] = (float(row[7]), float(row[8]))
                    except (TypeError, ValueError):
                        pass
    if lighting_info_dir and Path(lighting_info_dir).is_dir():
        li = Path(lighting_info_dir)
        d65_files = sorted((li / "D65 ").glob("ESPD_D65RING*.xls")) or sorted((li / "D65 ").glob("ESPD_D65*.xls"))
        f12_files = sorted((li / "Illuminant F1").glob("ESPD_F12*.xls"))
        if d65_files:
            meta, _, _ = parse_spectrum_file(d65_files[-1])
            if "x" in meta and "y" in meta:
                out["D65"] = (float(meta["x"]), float(meta["y"]))
        if f12_files:
            xs, ys = [], []
            for f in f12_files:
                meta, _, _ = parse_spectrum_file(f)
                if "x" in meta and "y" in meta:
                    xs.append(float(meta["x"]))
                    ys.append(float(meta["y"]))
            if xs:
                out["F12"] = (float(np.mean(xs)), float(np.mean(ys)))
    return out


def default_lighting_info_dir(data_root: Optional[Path] = None) -> Optional[Path]:
    root = Path(data_root or default_data_root())
    hit = root / "Lighting Information"
    return hit if hit.is_dir() else None


def default_torch_dir() -> Path:
    for pat in ("Torch_meas*/Torch_meas", "Torch_meas"):
        hit = resolve_downloads_path(pat)
        if hit and list(hit.glob("ESPD_T*.xls")):
            return hit
    return Path("/home/mabl-main/Downloads/Torch_meas-20260829T192551Z-1-001 (2)/Torch_meas")


def load_booth_fitskin_labs(xlsx: Path) -> Dict[str, Dict[str, np.ndarray]]:
    """person -> {D65|F12: Lab ndarray}."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    out: Dict[str, Dict[str, np.ndarray]] = {}
    sheet_map = {"Linh": "Lihn", "Lihn": "Lihn", "Parker": "Parker", "Anjana": "Anjana", "Woojae": "Woojae"}
    for sn in wb.sheetnames:
        person = sheet_map.get(sn.strip())
        if person is None:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        labs: Dict[str, np.ndarray] = {}
        for i, row in enumerate(rows):
            if row is None:
                continue
            # Pattern: row with D65/F12 label in col 6, Lab in cols 7-9
            if len(row) >= 10 and row[6] in ("D65", "F12"):
                try:
                    L, a, b = float(row[7]), float(row[8]), float(row[9])
                    labs[str(row[6])] = np.array([L, a, b], dtype=np.float64)
                except (TypeError, ValueError):
                    pass
            # Alternate: header row with L*,a*,b* then next rows
            if (
                len(row) >= 10
                and str(row[7]).strip() == "L*"
                and i + 1 < len(rows)
            ):
                for j in range(i + 1, min(i + 4, len(rows))):
                    r2 = rows[j]
                    if r2[6] in ("D65", "F12") and r2[7] is not None:
                        try:
                            labs[str(r2[6])] = np.array(
                                [float(r2[7]), float(r2[8]), float(r2[9])], dtype=np.float64
                            )
                        except (TypeError, ValueError):
                            pass
        if labs:
            out[person] = labs
    if not out:
        raise ValueError(f"No FitSkin Labs parsed from {xlsx}")
    return out


def normalize_person(name: str) -> Optional[str]:
    key = re.sub(r"[^a-z]", "", name.lower())
    return PERSON_ALIASES.get(key)


def parse_zip_stem(stem: str, folder_illuminant: str) -> Optional[Dict[str, Any]]:
    s = stem.replace(" ", "")
    patterns = [
        r"^(?P<person>[A-Za-z]+)[\-_]?(?P<ill>D65|F12|D12)[\-_]?(?P<cell>[A-E])(?P<rep>\d+)Torch$",
        r"^(?P<person>[A-Za-z]+)(?P<ill>F12|D65)(?P<cell>[A-E])(?P<rep>\d+)Torch$",
        r"^(?P<person>[A-Za-z]+)[—\-](?P<ill>D65|F12)[—\-](?P<cell>[A-E])(?P<rep>\d+)Torch$",
    ]
    m = None
    for pat in patterns:
        m = re.match(pat, s, re.I)
        if m:
            break
    if not m:
        return None
    person = normalize_person(m.group("person"))
    if person is None:
        return None
    ill = m.group("ill").upper()
    if ill == "D12":
        ill = "F12"
    cell = m.group("cell").upper()
    rep = int(m.group("rep"))
    # Trust filename illuminant over folder when they disagree (typos in folder layout)
    folder_ill = folder_illuminant.upper()
    if folder_ill == "D12":
        folder_ill = "F12"
    illuminant = ill if ill in ORACLE_CCT else folder_ill
    return {
        "person": person,
        "illuminant": illuminant,
        "wb_cell": cell,
        "rep": rep,
        "subject_id": f"{person}_{illuminant}_{cell}{rep}",
    }


def discover_trials(data_root: Path) -> List[Dict[str, Any]]:
    trials: List[Dict[str, Any]] = []
    for person_dir in sorted(data_root.iterdir()):
        if not person_dir.is_dir() or " " not in person_dir.name:
            continue
        for ill_dir_name in ("D65", "F12"):
            ill_dir = person_dir / ill_dir_name
            if not ill_dir.is_dir():
                continue
            for zpath in sorted(ill_dir.glob("*.zip")):
                if "torch" not in zpath.stem.lower():
                    continue
                meta = parse_zip_stem(zpath.stem, ill_dir_name)
                if meta is None:
                    print(f"warn: skip unparsed {zpath.name}")
                    continue
                meta["zip_path"] = zpath
                meta["zip_stem"] = zpath.stem
                meta["folder_illuminant"] = ill_dir_name
                trials.append(meta)
    return trials


def process_trial(
    trial: Dict[str, Any],
    *,
    M: np.ndarray,
    fairface: FairFacePredictor,
    torch_prior: TorchPrior,
    fitskin: np.ndarray,
    ring_xy: Dict[str, Tuple[float, float]],
    half_size: bool,
    lab_affine_W: Optional[np.ndarray] = None,
    multi_lab_bundle: Optional[Dict[str, Any]] = None,
    color_projector: Optional[Any] = None,
    M_default: Optional[np.ndarray] = None,
    M_warm: Optional[np.ndarray] = None,
    M_d65_ring: Optional[np.ndarray] = None,
) -> Dict[str, Any]:
    tmp = Path(tempfile.mkdtemp(prefix="ring_torch_"))
    try:
        nf, fl, lm_path = extract_zip(trial["zip_path"], tmp)
        A0 = load_dng_linear(nf, half_size=half_size, use_camera_wb=False)
        B0 = load_dng_linear(fl, half_size=half_size, use_camera_wb=False)
        if B0.shape != A0.shape:
            B0 = cv2.resize(B0, (A0.shape[1], A0.shape[0]), interpolation=cv2.INTER_AREA)
        lm = load_apple_landmarks(lm_path)
        skin_mask = apple_face_skin_roi_mask(
            lm, A0.shape[0], A0.shape[1], roi="forehead", linear_rgb=A0
        )
        if int(np.count_nonzero(skin_mask)) < 50:
            raise RuntimeError("empty forehead mask")

        preview = linear_rgb_to_preview_bgr(A0)
        face_rgb = face_rgb_crop_from_landmarks(preview, lm, padding=0.35)
        ff = fairface.predict_rgb(face_rgb)
        ethnicity = ff["predicted_ethnicity"]

        B0m, flash_scale = match_flash_exposure(A0, B0, skin_mask)
        R0 = np.sqrt(np.maximum(A0, 0) * np.maximum(B0m, 0) + 1e-8)
        cc_white_scale = in_frame_cc_white_scale(R0, preview)

        align_plain = align_flash_linear(A0, B0, cheek_mask=skin_mask, use_ecc=False)
        align_ecc = align_flash_linear(A0, B0, cheek_mask=skin_mask, use_ecc=True)

        lu_cct_res = estimate_lu(align_plain, torch_prior, use_measured_spd_rgb=False)
        lu_spd_res = estimate_lu(align_ecc, torch_prior, use_measured_spd_rgb=True)

        lu_cct = float(lu_cct_res.ambient_cct_estimated_k)
        lu_spd_cct = float(lu_spd_res.ambient_cct_estimated_k)

        nf_pix = A0[skin_mask > 0]
        nf_median = np.median(nf_pix, axis=0)
        nf_cct = chroma_to_planck_cct(np.maximum(nf_median, 1e-8))
        fused_cct = fused_lu_nf_cct(lu_spd_cct, nf_median, lu_weight=0.8)
        hybrid_cct = hybrid_deploy_cct(trial["illuminant"], lu_spd_cct)

        ill = trial["illuminant"]
        mk_xy = ring_xy.get(ill, ring_xy.get("F12" if ill == "D12" else "D65", DEFAULT_RING_XY["D65"]))
        illum_err = illuminant_vs_mk350(lu_spd_cct, mk_xy)

        cat_specs: Dict[str, Dict[str, Any]] = {
            "frozen_5500": {"kind": "mode", "mode": "frozen_5500"},
            "lu_cct": {"kind": "cct", "cct": lu_cct},
            "lu_spd_ecc": {"kind": "cct", "cct": lu_spd_cct},
            "lu_fused": {"kind": "cct", "cct": fused_cct},
            "hybrid_deploy": {"kind": "cct", "cct": hybrid_cct},
            "hybrid_lab_affine": {"kind": "cct", "cct": hybrid_cct, "lab_affine": True},
            "hybrid_color_projector": {"kind": "cct", "cct": hybrid_cct, "projector": True},
            "hybrid_multi_lab": {"kind": "cct", "cct": hybrid_cct, "multi_lab": True, "multi_lab_loo": True},
            "hybrid_multi_lab_allfit": {"kind": "cct", "cct": hybrid_cct, "multi_lab": True},
            # CC-supervised affines fit R₀→D65 XYZ directly; extra CAT double-corrects.
            "hybrid_routed_affine": {"kind": "identity", "routed_affine": True},
            "hybrid_routed_affine_multi_lab": {
                "kind": "identity",
                "routed_affine": True,
                "multi_lab": True,
                "multi_lab_loo": True,
            },
            "mk350_ring_xy": {"kind": "xy", "xy": mk_xy},
        }

        M_base = M_default if M_default is not None else M
        _, default_affine_tag = select_affine_for_illuminant(
            ill, M_default=M_base, M_warm=M_warm, M_d65_ring=M_d65_ring
        )

        row: Dict[str, Any] = {
            **trial,
            "zip_path": str(trial["zip_path"]),
            "affine_route": default_affine_tag,
            "fitskin_L": float(fitskin[0]),
            "fitskin_a": float(fitskin[1]),
            "fitskin_b": float(fitskin[2]),
            "flash_scale": float(flash_scale),
            "ecc_cc": float(align_ecc.ecc_cc),
            "torch_cct_prior_k": float(torch_prior.torch_cct_k),
            "lu_cct_k": lu_cct,
            "lu_spd_ecc_k": lu_spd_cct,
            "lu_fused_cct_k": fused_cct,
            "hybrid_cct_k": hybrid_cct,
            "nf_chroma_cct_k": float(nf_cct),
            "mk350_ring_x": float(mk_xy[0]),
            "mk350_ring_y": float(mk_xy[1]),
            "mk350_cct_k": illum_err["mk350_cct_k"],
            "lu_spd_delta_cct_k": illum_err["delta_cct_k"],
            "lu_spd_abs_delta_cct_k": illum_err["abs_delta_cct_k"],
            "fairface_ethnicity": ethnicity,
            "fairface_confidence": float(ff["confidence"]),
            "cc_white_scale": float(cc_white_scale) if cc_white_scale is not None else np.nan,
            "cc_detected": cc_white_scale is not None,
        }

        R0_cc = R0 * cc_white_scale if cc_white_scale is not None else R0

        for arm, spec in cat_specs.items():
            kind = spec["kind"]
            if kind == "xy":
                val = spec["xy"]
                xyz_w = xy_white_y1(val[0], val[1])
                row[f"cat_cct_{arm}"] = xy_to_cct_mccamy(val[0], val[1])
                row[f"cat_xy_{arm}"] = f"{val[0]:.4f},{val[1]:.4f}"
            elif kind == "mode":
                xyz_w, cat_meta = scene_white_for_cat(
                    spec["mode"],
                    illuminant_label=ill,
                    lu_cct=lu_spd_cct,
                    mk350_xy=mk_xy,
                )
                row[f"cat_cct_{arm}"] = cat_meta["cat_cct"]
                row[f"cat_xy_{arm}"] = ""
            elif kind == "identity":
                xyz_w = None
                row[f"cat_cct_{arm}"] = np.nan
                row[f"cat_xy_{arm}"] = "identity"
            else:
                xyz_w = planck_xyz_y1(float(spec["cct"]))
                row[f"cat_cct_{arm}"] = float(spec["cct"])
                row[f"cat_xy_{arm}"] = ""

            projector = color_projector if spec.get("projector") else None
            M_use, affine_tag = (
                select_affine_for_illuminant(
                    ill, M_default=M_base, M_warm=M_warm, M_d65_ring=M_d65_ring
                )
                if spec.get("routed_affine")
                else (M, "tier3_default")
            )
            if spec.get("routed_affine"):
                row[f"affine_tag_{arm}"] = affine_tag
            cat_degree = 0.0 if spec.get("routed_affine") else 1.0
            rgb_in = R0_cc if spec.get("routed_affine") else R0
            lab, _ = mean_lab_on_mask(
                rgb_in,
                skin_mask,
                M_use,
                projector=projector,
                xyz_scene_white=xyz_w,
                cat_degree=cat_degree,
                l_sampling="off",
                ethnicity=ethnicity,
                skin_lab_trim=FOREHEAD_SKIN_LAB_TRIM,
            )
            if spec.get("lab_affine") and lab_affine_W is not None:
                lab = apply_lab_affine(lab, lab_affine_W)
            if spec.get("multi_lab") and multi_lab_bundle is not None:
                lab, _ = apply_multi_illuminant_lab_affine(
                    lab,
                    ill,
                    multi_lab_bundle,
                    mode="routed",
                    person_key=f"ring_{trial['person']}",
                    loo=bool(spec.get("multi_lab_loo")),
                )
            row[f"pred_L_{arm}"] = float(lab[0])
            row[f"pred_a_{arm}"] = float(lab[1])
            row[f"pred_b_{arm}"] = float(lab[2])
            row[f"de00_{arm}"] = float(delta_e_2000(lab, fitskin))

        row["delta_f12_lu_vs_frozen"] = (
            row["de00_frozen_5500"] - row["de00_lu_spd_ecc"] if ill == "F12" else np.nan
        )
        return row
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def summarize(rows: List[Dict[str, Any]], arms: List[str]) -> Dict[str, Any]:
    def stats(vals: List[float]) -> Dict[str, Any]:
        v = [float(x) for x in vals if x is not None and np.isfinite(x)]
        return {"n": len(v), "mean": round(mean(v), 4) if v else None, "median": round(median(v), 4) if v else None}

    out: Dict[str, Any] = {
        "n": len(rows),
        "overall": {},
        "by_illuminant": {},
        "by_person_illuminant": {},
        "by_wb_cell": {},
        "by_illuminant_wb_cell": {},
        "illuminant_estimation": {},
    }
    for arm in arms:
        out["overall"][arm] = stats([r[f"de00_{arm}"] for r in rows])
    by_ill: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_pi: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_wb: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_ill_wb: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_ill[r["illuminant"]].append(r)
        by_pi[f"{r['person']}_{r['illuminant']}"].append(r)
        cell = str(r.get("wb_cell") or "").strip().upper()
        if cell:
            by_wb[cell].append(r)
            by_ill_wb[f"{r['illuminant']}_{cell}"].append(r)
    for ill, grp in sorted(by_ill.items()):
        out["by_illuminant"][ill] = {arm: stats([x[f"de00_{arm}"] for x in grp]) for arm in arms}
    for key, grp in sorted(by_pi.items()):
        out["by_person_illuminant"][key] = {arm: stats([x[f"de00_{arm}"] for x in grp]) for arm in arms}
    for cell, grp in sorted(by_wb.items()):
        out["by_wb_cell"][cell] = {arm: stats([x[f"de00_{arm}"] for x in grp]) for arm in arms}
    for key, grp in sorted(by_ill_wb.items()):
        ill, cell = key.split("_", 1)
        out["by_illuminant_wb_cell"].setdefault(ill, {})[cell] = {
            arm: stats([x[f"de00_{arm}"] for x in grp]) for arm in arms
        }

    out["illuminant_estimation"] = {
        "lu_spd_abs_delta_cct_k": stats([r["lu_spd_abs_delta_cct_k"] for r in rows]),
        "lu_spd_delta_cct_k": stats([r["lu_spd_delta_cct_k"] for r in rows]),
        "ecc_cc": stats([r["ecc_cc"] for r in rows]),
    }
    for ill in ("D65", "F12"):
        grp = [r for r in rows if r["illuminant"] == ill]
        if grp:
            out["illuminant_estimation"][f"{ill}_lu_spd_abs_delta_cct_k"] = stats(
                [r["lu_spd_abs_delta_cct_k"] for r in grp]
            )

    for label, arm in (
        ("f12_lu_spd_ecc_improvement_vs_frozen", "lu_spd_ecc"),
        ("f12_hybrid_improvement_vs_frozen", "hybrid_deploy"),
        ("f12_lu_fused_improvement_vs_frozen", "lu_fused"),
    ):
        f12 = [r for r in rows if r["illuminant"] == "F12"]
        if f12:
            out[label] = stats([r["de00_frozen_5500"] - r[f"de00_{arm}"] for r in f12])
    return out


def write_illuminant_plots(rows: List[Dict[str, Any]], out_dir: Path) -> None:
    """Lu SPD+ECC CCT vs MK350 in-situ ring CCT scatter."""
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        return
    if not rows:
        return
    mk = np.array([r["mk350_cct_k"] for r in rows], dtype=float)
    lu = np.array([r["lu_spd_ecc_k"] for r in rows], dtype=float)
    ills = [r["illuminant"] for r in rows]
    fig, ax = plt.subplots(figsize=(5.5, 5.5))
    colors = {"D65": "#2563eb", "F12": "#dc2626"}
    for ill in sorted(set(ills)):
        m = np.array([x == ill for x in ills])
        ax.scatter(mk[m], lu[m], s=28, alpha=0.75, label=ill, c=colors.get(ill, "#666"))
    lo = float(min(mk.min(), lu.min()) - 200)
    hi = float(max(mk.max(), lu.max()) + 200)
    ax.plot([lo, hi], [lo, hi], "k--", lw=1, alpha=0.5)
    ax.set_xlabel("MK350 ring CCT (K)")
    ax.set_ylabel("Lu SPD+ECC estimated CCT (K)")
    ax.set_title("Illuminant estimation vs in-situ MK350")
    ax.legend()
    ax.set_aspect("equal", adjustable="box")
    fig.tight_layout()
    fig.savefig(out_dir / "lu_vs_mk350_cct.png", dpi=150)
    plt.close(fig)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--data-root", type=Path, default=None, help="Ring-light dataset root (auto-detected from ~/Downloads)")
    ap.add_argument("--booth-xlsx", type=Path, default=Path.home() / "Downloads" / "Booth Lighting.xlsx")
    ap.add_argument("--torch-dir", type=Path, default=None, help="MK350 Torch_meas folder (auto-detected from ~/Downloads)")
    ap.add_argument("--out-dir", type=Path, default=ROOT / "results" / "torch_illuminant_ringlight")
    ap.add_argument("--cal-dir", type=Path, default=ROOT / "calibration" / "tier3_affine")
    ap.add_argument("--fairface-dir", type=Path, default=ROOT / "calibration" / "fairface")
    ap.add_argument(
        "--lab-corrector",
        type=Path,
        default=ROOT / "calibration" / "lab_affine_corrector_pansor" / "lab_affine_4x3.npy",
    )
    ap.add_argument(
        "--multi-lab-corrector",
        type=Path,
        default=ROOT / "calibration" / "multi_illuminant_lab_affine" / "multi_illuminant_lab_affine.json",
    )
    ap.add_argument(
        "--color-projector",
        type=Path,
        default=ROOT / "calibration" / "color_projector_pansor" / "color_projector.npz",
    )
    ap.add_argument(
        "--warm-cal-dir",
        type=Path,
        default=ROOT / "calibration" / "tier3_affine_warm_cc",
    )
    ap.add_argument(
        "--d65-ring-cal-dir",
        type=Path,
        default=ROOT / "calibration" / "tier3_affine_d65_ring_cc",
    )
    ap.add_argument("--half-size", action="store_true", default=True)
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    data_root = Path(args.data_root or default_data_root()).expanduser().resolve()
    torch_dir = Path(args.torch_dir or default_torch_dir()).expanduser().resolve()
    booth_xlsx = Path(args.booth_xlsx).expanduser().resolve()

    arms = list(PRIMARY_ARMS)
    fitskin_map = load_booth_fitskin_labs(booth_xlsx)
    torch_prior = load_torch_prior(torch_dir)
    ring_xy = load_ring_illuminant_xy(
        booth_xlsx=booth_xlsx,
        lighting_info_dir=default_lighting_info_dir(data_root),
    )
    trials = discover_trials(data_root)
    print(f"data-root: {data_root}")
    print(f"torch-dir: {torch_dir}  (CCT prior {torch_prior.torch_cct_k:.0f} K)")
    print(f"booth-xlsx: {booth_xlsx}")
    print(f"MK350 ring xy: D65={ring_xy['D65']}  F12={ring_xy['F12']}")
    print(f"trials: {len(trials)}")
    if args.limit > 0:
        trials = trials[: args.limit]

    M = load_affine(args.cal_dir)
    M_warm: Optional[np.ndarray] = None
    M_d65_ring: Optional[np.ndarray] = None
    warm_npy = Path(args.warm_cal_dir) / "camera_rgb_to_xyz_affine.npy"
    d65_npy = Path(args.d65_ring_cal_dir) / "camera_rgb_to_xyz_affine.npy"
    if warm_npy.is_file():
        M_warm = np.load(warm_npy)
    if d65_npy.is_file():
        M_d65_ring = np.load(d65_npy)
    fairface = FairFacePredictor.load(mode="7", weights_dir=args.fairface_dir)

    lab_affine_W: Optional[np.ndarray] = None
    if args.lab_corrector.is_file():
        lab_affine_W = load_lab_affine(args.lab_corrector)
    multi_lab_bundle: Optional[Dict[str, Any]] = None
    if args.multi_lab_corrector.is_file():
        multi_lab_bundle = load_multi_illuminant_lab_affine(args.multi_lab_corrector)
    color_projector = None
    if args.color_projector.is_file():
        color_projector = load_color_projector(args.color_projector)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, str]] = []

    for i, trial in enumerate(trials, 1):
        person = trial["person"]
        ill = trial["illuminant"]
        if person not in fitskin_map or ill not in fitskin_map[person]:
            errors.append({"zip": str(trial["zip_path"]), "error": f"no FitSkin GT for {person}/{ill}"})
            continue
        try:
            row = process_trial(
                trial,
                M=M,
                fairface=fairface,
                torch_prior=torch_prior,
                fitskin=fitskin_map[person][ill],
                ring_xy=ring_xy,
                half_size=bool(args.half_size),
                lab_affine_W=lab_affine_W,
                multi_lab_bundle=multi_lab_bundle,
                color_projector=color_projector,
                M_default=M,
                M_warm=M_warm,
                M_d65_ring=M_d65_ring,
            )
            rows.append(row)
            if i % 10 == 0:
                print(f"  [{i}/{len(trials)}] ok", trial["subject_id"])
        except Exception as exc:
            errors.append({"zip": str(trial["zip_path"]), "error": str(exc)})
            print(f"  FAIL {trial['zip_stem']}: {exc}")

    summary = summarize(rows, arms)
    summary["torch_prior"] = {
        "torch_cct_k": torch_prior.torch_cct_k,
        "n_files": torch_prior.n_files,
        "files": list(torch_prior.files),
    }
    summary["oracle_cct"] = ORACLE_CCT
    summary["mk350_ring_xy"] = {k: list(v) for k, v in ring_xy.items()}
    summary["n_fail"] = len(errors)
    summary["fitskin"] = {p: {k: v.tolist() for k, v in d.items()} for p, d in fitskin_map.items()}
    summary["note"] = (
        "Lu uses flash/no-flash + MK350 torch SPD prior; hybrid_deploy: Lu on F12/warm, frozen 5500 on D65. "
        "Stacked arms: hybrid + lab affine, color projector, or multi-illuminant lab corrector."
    )
    summary["stacked_correctors"] = {
        "lab_corrector": str(args.lab_corrector) if lab_affine_W is not None else None,
        "multi_lab_corrector": str(args.multi_lab_corrector) if multi_lab_bundle is not None else None,
        "color_projector": str(args.color_projector) if color_projector is not None else None,
        "warm_cal_dir": str(args.warm_cal_dir) if M_warm is not None else None,
        "d65_ring_cal_dir": str(args.d65_ring_cal_dir) if M_d65_ring is not None else None,
    }

    csv_path = args.out_dir / "torch_illuminant_ringlight.csv"
    if rows:
        fieldnames = list(rows[0].keys())
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    tsv_path = args.out_dir / "torch_illuminant_for_sheets.tsv"
    with tsv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(
            [
                "subject_id",
                "person",
                "illuminant",
                "wb_cell",
                "lu_cct_k",
                "lu_spd_ecc_k",
                "lu_fused_cct_k",
                "hybrid_cct_k",
                "mk350_cct_k",
                "lu_spd_abs_delta_cct_k",
                "de00_frozen_5500",
                "de00_lu_cct",
                "de00_lu_spd_ecc",
                "de00_lu_fused",
                "de00_hybrid_deploy",
                "de00_hybrid_lab_affine",
                "de00_hybrid_color_projector",
                "de00_hybrid_multi_lab",
                "de00_mk350_ring_xy",
                "delta_de00_lu_spd_vs_frozen",
            ]
        )
        for r in rows:
            w.writerow(
                [
                    r["subject_id"],
                    r["person"],
                    r["illuminant"],
                    r["wb_cell"],
                    round(r["lu_cct_k"], 1),
                    round(r["lu_spd_ecc_k"], 1),
                    round(r["lu_fused_cct_k"], 1),
                    round(r["hybrid_cct_k"], 1),
                    round(r["mk350_cct_k"], 1),
                    round(r["lu_spd_abs_delta_cct_k"], 1),
                    round(r["de00_frozen_5500"], 3),
                    round(r["de00_lu_cct"], 3),
                    round(r["de00_lu_spd_ecc"], 3),
                    round(r["de00_lu_fused"], 3),
                    round(r["de00_hybrid_deploy"], 3),
                    round(r.get("de00_hybrid_lab_affine", float("nan")), 3),
                    round(r.get("de00_hybrid_color_projector", float("nan")), 3),
                    round(r.get("de00_hybrid_multi_lab", float("nan")), 3),
                    round(r["de00_mk350_ring_xy"], 3),
                    round(r["de00_frozen_5500"] - r["de00_lu_spd_ecc"], 3),
                ]
            )

    write_illuminant_plots(rows, args.out_dir)
    (args.out_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    if errors:
        (args.out_dir / "errors.json").write_text(json.dumps(errors, indent=2) + "\n", encoding="utf-8")

    if len(rows) >= 80:
        from scripts.summarize_ring_eval_by_wb_cell import summarize_rows as wb_summarize

        wb_out = ROOT / "data" / "ring_light" / "eval_n84_by_wb_cell.json"
        wb_payload = wb_summarize(rows, ["frozen_5500", "hybrid_deploy", "hybrid_multi_lab"])
        wb_payload["source_csv"] = str(csv_path.resolve())
        wb_payload["source_note"] = (
            "Forehead ROI, pre-AWB tier3 affine, hybrid_deploy + multi-lab (evaluate_ringlight_torch_illuminant)"
        )
        wb_out.parent.mkdir(parents=True, exist_ok=True)
        wb_out.write_text(json.dumps(wb_payload, indent=2) + "\n", encoding="utf-8")
        print(f"Wrote {wb_out}")

    print(json.dumps(summary["overall"], indent=2))
    print(f"\nWrote {csv_path} ({len(rows)} rows, {len(errors)} errors)")


if __name__ == "__main__":
    main()
