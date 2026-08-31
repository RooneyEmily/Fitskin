#!/usr/bin/env python3
"""Evaluate Pansor CameraSettingsDataset: ΔE00 vs shutter / ISO sweeps.

Each zip is flash/no-flash + Apple landmarks (same layout as Pansor indoor).
FitSkin Inside Lab and nominal camera settings come from
``Camera Settings Aug10.xlsx`` (one sheet per person).

Emily has Lab + planned settings in the spreadsheet but **no capture zips**.

Primary color path (claimable):
  --scr-mode preawb_cat --fixed-cat-k 5500 --l-sampling off
Also reports FairFace-7 ROI sampling for comparison.

Example:
  python3 scripts/evaluate_camera_settings.py \\
    --data-root ~/Downloads/CameraSettingsDataset-.../CameraSettingsDataset \\
    --out-dir results/camera_settings
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from statistics import mean, median
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from delta_e_2000 import delta_e_2000  # noqa: E402
from flash_noflash_spectral import planck_xyz_y1  # noqa: E402
from scripts.evaluate_pansor20_chartfree_d65 import (  # noqa: E402
    apple_face_cheek_masks,
    extract_zip,
    load_affine,
    load_apple_landmarks,
    load_dng_linear,
    match_flash_exposure,
    mean_lab_on_mask,
)

# Setting letter → role in the factorial design (from spreadsheet).
# A–D: shutter sweep (ISO≈fixed). E–H: ISO sweep (shutter≈fixed). I: Wooj-only near-baseline ISO.
SHUTTER_SETTINGS = ("A", "B", "C", "D")
ISO_SETTINGS = ("E", "F", "G", "H")


def parse_shutter_to_seconds(s: str) -> float:
    """'1/127s' → 1/127."""
    s = str(s).strip().lower().replace(" ", "")
    s = s.rstrip("s")
    if "/" in s:
        a, b = s.split("/", 1)
        return float(a) / float(b)
    return float(s)


def parse_wb_k(s: str) -> float:
    return float(str(s).strip().upper().replace("K", ""))


def load_settings_workbook(xlsx: Path) -> Dict[str, Dict[str, Any]]:
    """person -> {fitskin_Lab, settings: {letter: {shutter_s, iso, wb_k, label}}}."""
    wb = load_workbook(xlsx, data_only=True)
    out: Dict[str, Dict[str, Any]] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        person = sn.strip()
        if person.lower().startswith("gian"):
            person = "Giana"
        elif person.lower().startswith("wooj"):
            person = "Wooj"
        settings: Dict[str, Dict[str, Any]] = {}
        fitskin = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            # Lab footer rows start with numeric L
            if isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
                # Could be Lab row if a,b present and Name empty-ish
                if row[1] is not None and row[2] is not None and (
                    row[3] is None or str(row[3]).strip() == ""
                ):
                    fitskin = {
                        "fitskin_L": float(row[0]),
                        "fitskin_a": float(row[1]),
                        "fitskin_b": float(row[2]),
                    }
                continue
            shutter, iso, wb_cell, name = row[0], row[1], row[2], row[3]
            if name is None or iso is None:
                continue
            name_s = str(name)
            # e.g. 'Wooj A1,2,3' or 'Giana A1,2,3'
            m = re.search(r"\b([A-I])\d", name_s, re.I)
            if not m:
                continue
            letter = m.group(1).upper()
            settings[letter] = {
                "shutter_s": parse_shutter_to_seconds(shutter),
                "iso": float(iso),
                "wb_k": parse_wb_k(wb_cell),
                "label": name_s,
                "shutter_raw": str(shutter),
            }
        # Lab is on rows labeled L,a,b then values — catch that pattern
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if row and str(row[0]).strip().upper() == "L" and i + 1 < len(rows):
                lab = rows[i + 1]
                if lab and lab[0] is not None:
                    fitskin = {
                        "fitskin_L": float(lab[0]),
                        "fitskin_a": float(lab[1]),
                        "fitskin_b": float(lab[2]),
                    }
        if not settings or fitskin is None:
            print(f"warn: incomplete sheet {sn!r} settings={len(settings)} lab={fitskin}")
            continue
        out[person] = {"fitskin": fitskin, "settings": settings}
    return out


def parse_zip_name(stem: str) -> Optional[Tuple[str, str, int]]:
    m = re.match(r"(Giana|Gianna|Keaton|Parker|Wooj|Emily)([A-I])(\d+)$", stem, re.I)
    if not m:
        return None
    person, letter, rep = m.group(1), m.group(2).upper(), int(m.group(3))
    if person.lower().startswith("gian"):
        person = "Giana"
    elif person.lower().startswith("wooj"):
        person = "Wooj"
    else:
        person = person.capitalize() if person.lower() != "emily" else "Emily"
        if person.lower() == "keaton":
            person = "Keaton"
        if person.lower() == "parker":
            person = "Parker"
    return person, letter, rep


def discover_trials(data_root: Path) -> List[Dict[str, Any]]:
    rows = []
    for zpath in sorted(data_root.glob("*.zip")):
        parsed = parse_zip_name(zpath.stem)
        if parsed is None:
            print(f"skip unparsed zip: {zpath.name}")
            continue
        person, letter, rep = parsed
        rows.append(
            {
                "person": person,
                "setting": letter,
                "rep": rep,
                "zip_path": zpath,
                "zip_stem": zpath.stem,
                "subject_id": f"{person}_{letter}{rep}",
            }
        )
    return rows


def process_one(
    trial: Dict[str, Any],
    *,
    work_dir: Path,
    M: np.ndarray,
    xyz_w: np.ndarray,
    fit: np.ndarray,
    half_size: bool,
    fairface_predictor: Any,
) -> Dict[str, Any]:
    nf, fl, lm_path = extract_zip(trial["zip_path"], work_dir / trial["subject_id"])
    A0 = load_dng_linear(nf, half_size=half_size, use_camera_wb=False)
    B0 = load_dng_linear(fl, half_size=half_size, use_camera_wb=False)
    if B0.shape != A0.shape:
        B0 = cv2.resize(B0, (A0.shape[1], A0.shape[0]), interpolation=cv2.INTER_AREA)
    lm = load_apple_landmarks(lm_path)
    _, cheek = apple_face_cheek_masks(lm, A0.shape[0], A0.shape[1])
    n_cheek = int(np.count_nonzero(cheek))
    if n_cheek < 50:
        raise RuntimeError(f"empty cheek ({n_cheek})")

    B0m, flash_scale = match_flash_exposure(A0, B0, cheek)
    R0 = np.sqrt(np.maximum(A0, 0) * np.maximum(B0m, 0) + 1e-8)

    Lab_off, _ = mean_lab_on_mask(
        R0, cheek, M, xyz_scene_white=xyz_w, cat_degree=1.0, l_sampling="off"
    )
    de_off = float(delta_e_2000(Lab_off, fit))

    out: Dict[str, Any] = {
        **{k: trial[k] for k in ("person", "setting", "rep", "zip_stem", "subject_id")},
        "zip_path": str(trial["zip_path"]),
        "n_cheek": n_cheek,
        "flash_scale": float(flash_scale),
        "pipeline_L_off": float(Lab_off[0]),
        "pipeline_a_off": float(Lab_off[1]),
        "pipeline_b_off": float(Lab_off[2]),
        "de00_off": de_off,
        "fitskin_L": float(fit[0]),
        "fitskin_a": float(fit[1]),
        "fitskin_b": float(fit[2]),
    }

    if fairface_predictor is not None:
        from scripts.evaluate_pansor20_chartfree_d65 import linear_rgb_to_preview_bgr
        from models.fairface_race import face_rgb_crop_from_landmarks

        preview = linear_rgb_to_preview_bgr(A0)
        face_rgb = face_rgb_crop_from_landmarks(preview, lm, padding=0.35)
        ff = fairface_predictor.predict_rgb(face_rgb)
        Lab_ff, sm = mean_lab_on_mask(
            R0,
            cheek,
            M,
            xyz_scene_white=xyz_w,
            cat_degree=1.0,
            l_sampling="specular_tone",
            ethnicity=ff["predicted_ethnicity"],
        )
        out.update(
            {
                "fairface_label": ff["fairface_label"],
                "predicted_ethnicity": ff["predicted_ethnicity"],
                "fairface_confidence": float(ff["confidence"]),
                "pipeline_L_ff": float(Lab_ff[0]),
                "pipeline_a_ff": float(Lab_ff[1]),
                "pipeline_b_ff": float(Lab_ff[2]),
                "de00_ff": float(delta_e_2000(Lab_ff, fit)),
                "ff_sampling": json.dumps(sm),
            }
        )
    return out


# Okabe–Ito–style palette (no red/green neighbors in legend order).
_PERSON_COLORS = {
    "Giana": "#0072B2",   # blue
    "Keaton": "#E69F00",  # amber
    "Parker": "#56B4E9",  # sky
    "Wooj": "#CC79A7",    # pink
}
_REF_COLOR = "#666666"  # neutral gray reference (not green)
_LINE_SHUTTER = "#0072B2"
_LINE_ISO = "#D55E00"  # vermillion; ref is gray, not green


def _agg_by_x(
    rows: List[Dict[str, Any]],
    letters: Tuple[str, ...],
    xkey: str,
    ykey: str = "de00_off",
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Aggregate mean±std per setting letter, then sort by x so the line is continuous."""
    pts = []
    for letter in letters:
        vals = [r for r in rows if r["setting"] == letter]
        if not vals:
            continue
        x = float(mean(r[xkey] for r in vals))
        yv = [float(r[ykey]) for r in vals]
        y = float(mean(yv))
        # Always report sample std (0 if n==1) so every point gets an errorbar artist.
        err = float(np.std(yv, ddof=0)) if len(yv) >= 1 else 0.0
        pts.append((x, y, err))
    pts.sort(key=lambda t: t[0])
    if not pts:
        return np.array([]), np.array([]), np.array([])
    xs, ys, yerr = zip(*pts)
    return np.asarray(xs), np.asarray(ys), np.asarray(yerr)


def make_plots(rows: List[Dict[str, Any]], out_dir: Path) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    people = sorted({r["person"] for r in rows})

    def _plot_sweep(
        letters: Tuple[str, ...],
        xkey: str,
        xlabel: str,
        title: str,
        out_name: str,
        color: str,
        marker: str,
    ) -> None:
        fig, axes = plt.subplots(1, len(people), figsize=(3.2 * len(people), 3.6), squeeze=False)
        for ax, person in zip(axes[0], people):
            sub = [r for r in rows if r["person"] == person and r["setting"] in letters]
            xs, ys, yerr = _agg_by_x(sub, letters, xkey=xkey, ykey="de00_off")
            if xs.size == 0:
                ax.set_title(f"{person} (no data)")
                continue
            # Sort-by-x so the polyline is continuous along the exposure axis.
            # Means only (no error bars): with n≈2–4, std whiskers look like
            # missing bars on low-variance cells and clutter the sweep shape.
            ax.plot(
                xs,
                ys,
                marker=marker,
                color=color,
                linewidth=1.5,
                markersize=6,
                zorder=3,
            )
            ax.set_xscale("log")
            ax.set_xlabel(xlabel)
            ax.set_ylabel(r"$\Delta E_{00}$ (frozen)")
            ax.set_title(person)
            ax.axhline(5.55, color=_REF_COLOR, ls="--", lw=1.0, alpha=0.9, zorder=1)
            ax.set_ylim(bottom=0)
        fig.suptitle(title, fontsize=12)
        fig.tight_layout()
        fig.savefig(out_dir / out_name, dpi=150)
        plt.close(fig)

    _plot_sweep(
        SHUTTER_SETTINGS,
        xkey="shutter_s",
        xlabel="Shutter (s)",
        title="Camera settings — shutter sweep (ISO≈fixed)",
        out_name="de00_vs_shutter.png",
        color=_LINE_SHUTTER,
        marker="o",
    )
    _plot_sweep(
        ISO_SETTINGS,
        xkey="iso",
        xlabel="ISO",
        title="Camera settings — ISO sweep (shutter≈fixed)",
        out_name="de00_vs_iso.png",
        color=_LINE_ISO,
        marker="s",
    )

    # --- Bar: mean ΔE by setting letter (A–H only; I is Wooj-only near-baseline) ---
    letters = [let for let in sorted({r["setting"] for r in rows}) if let != "I"]
    fig, ax = plt.subplots(figsize=(10, 4))
    x0 = np.arange(len(letters))
    w = 0.8 / max(len(people), 1)
    for i, person in enumerate(people):
        means = []
        for let in letters:
            vals = [r["de00_off"] for r in rows if r["person"] == person and r["setting"] == let]
            means.append(mean(vals) if vals else np.nan)
        ax.bar(
            x0 + i * w,
            means,
            w,
            label=person,
            color=_PERSON_COLORS.get(person, f"C{i}"),
        )
    ax.set_xticks(x0 + w * (len(people) - 1) / 2)
    ax.set_xticklabels(letters)
    ax.set_ylabel(r"mean $\Delta E_{00}$ (frozen)")
    ax.set_xlabel("Setting letter (A–D shutter · E–H ISO)")
    ax.axhline(5.55, color=_REF_COLOR, ls="--", lw=1.0, label="frozen ref 5.55")
    ax.legend(fontsize=8, frameon=False, ncol=2)
    ax.set_title("ΔE00 by camera-setting letter")
    fig.tight_layout()
    fig.savefig(out_dir / "de00_by_setting_letter.png", dpi=150)
    plt.close(fig)

    # L* vs setting — sort by x so lines don't zigzag
    fig, axes = plt.subplots(1, 2, figsize=(10, 3.8))
    for ax, letters_grp, title, xkey in [
        (axes[0], SHUTTER_SETTINGS, "Shutter sweep", "shutter_s"),
        (axes[1], ISO_SETTINGS, "ISO sweep", "iso"),
    ]:
        for person in people:
            sub = [r for r in rows if r["person"] == person and r["setting"] in letters_grp]
            xs, ys, _ = _agg_by_x(sub, letters_grp, xkey=xkey, ykey="pipeline_L_off")
            if xs.size:
                ax.plot(
                    xs,
                    ys,
                    marker="o",
                    label=person,
                    color=_PERSON_COLORS.get(person, None),
                    linewidth=1.5,
                )
        ax.set_xscale("log")
        ax.set_xlabel(xkey)
        ax.set_ylabel("Pipeline L* (frozen)")
        ax.set_title(title)
        ax.legend(fontsize=8, frameon=False)
    fig.suptitle("Cheek L* response to exposure settings", fontsize=12)
    fig.tight_layout()
    fig.savefig(out_dir / "Lstar_vs_settings.png", dpi=150)
    plt.close(fig)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--data-root",
        type=Path,
        default=Path(
            "/home/mabl-main/Downloads/CameraSettingsDataset-20260809T210945Z-1-001/CameraSettingsDataset"
        ),
    )
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Defaults to <data-root>/Camera Settings Aug10.xlsx",
    )
    ap.add_argument(
        "--cal-dir",
        type=Path,
        default=ROOT / "calibration" / "tier3_affine",
    )
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=ROOT / "results" / "camera_settings",
    )
    ap.add_argument("--work-dir", type=Path, default=None)
    ap.add_argument("--fixed-cat-k", type=float, default=5500.0)
    ap.add_argument("--full-res", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--person", type=str, default="", help="Only this person (e.g. Wooj)")
    ap.add_argument("--no-fairface", action="store_true")
    ap.add_argument(
        "--fairface-dir",
        type=Path,
        default=ROOT / "calibration" / "fairface",
    )
    args = ap.parse_args()

    xlsx = args.xlsx or (args.data_root / "Camera Settings Aug10.xlsx")
    work_dir = args.work_dir or (args.out_dir / "_extract")
    args.out_dir.mkdir(parents=True, exist_ok=True)
    work_dir.mkdir(parents=True, exist_ok=True)

    book = load_settings_workbook(xlsx)
    print("People in workbook:", sorted(book))
    for p, info in book.items():
        lab = info["fitskin"]
        print(
            f"  {p}: Lab=({lab['fitskin_L']:.2f},{lab['fitskin_a']:.2f},{lab['fitskin_b']:.2f}) "
            f"settings={''.join(sorted(info['settings']))}"
        )

    trials = discover_trials(args.data_root)
    if args.person:
        trials = [t for t in trials if t["person"].lower() == args.person.lower()]
    # Drop Emily / anyone without Lab+settings
    keep = []
    for t in trials:
        if t["person"] not in book:
            print(f"skip {t['zip_stem']}: no workbook entry for {t['person']}")
            continue
        if t["setting"] not in book[t["person"]]["settings"]:
            print(f"skip {t['zip_stem']}: setting {t['setting']} missing in workbook")
            continue
        keep.append(t)
    trials = keep
    if args.limit:
        trials = trials[: args.limit]
    print(f"Trials to process: {len(trials)}")

    M = load_affine(args.cal_dir)
    xyz_w = planck_xyz_y1(float(args.fixed_cat_k), 0.0)
    half_size = not args.full_res

    fairface_predictor = None
    if not args.no_fairface:
        from models.fairface_race import FairFacePredictor

        fairface_predictor = FairFacePredictor.load(mode="7", weights_dir=args.fairface_dir)
        print(f"FairFace-7 on {fairface_predictor.device}")

    rows: List[Dict[str, Any]] = []
    for i, t in enumerate(trials, 1):
        info = book[t["person"]]
        st = info["settings"][t["setting"]]
        fit = np.array(
            [
                info["fitskin"]["fitskin_L"],
                info["fitskin"]["fitskin_a"],
                info["fitskin"]["fitskin_b"],
            ],
            dtype=np.float64,
        )
        print(
            f"[{i}/{len(trials)}] {t['subject_id']:12s} shutter={st['shutter_raw']} "
            f"ISO={st['iso']:.0f} WB={st['wb_k']:.0f}K",
            flush=True,
        )
        try:
            r = process_one(
                t,
                work_dir=work_dir,
                M=M,
                xyz_w=xyz_w,
                fit=fit,
                half_size=half_size,
                fairface_predictor=fairface_predictor,
            )
            r.update(
                {
                    "shutter_s": st["shutter_s"],
                    "shutter_raw": st["shutter_raw"],
                    "iso": st["iso"],
                    "wb_k": st["wb_k"],
                    "setting_label": st["label"],
                }
            )
            rows.append(r)
            msg = f"  ΔE_off={r['de00_off']:.2f} L*={r['pipeline_L_off']:.1f}"
            if "de00_ff" in r:
                msg += f"  ΔE_ff={r['de00_ff']:.2f} ({r['predicted_ethnicity']})"
            print(msg, flush=True)
        except Exception as e:
            print(f"  FAIL: {e}", flush=True)

    if not rows:
        raise SystemExit("No successful trials")

    csv_path = args.out_dir / "camera_settings_results.csv"
    fields = sorted({k for r in rows for k in r})
    # Prefer a stable column order
    preferred = [
        "person",
        "setting",
        "rep",
        "subject_id",
        "shutter_raw",
        "shutter_s",
        "iso",
        "wb_k",
        "de00_off",
        "de00_ff",
        "pipeline_L_off",
        "pipeline_a_off",
        "pipeline_b_off",
        "pipeline_L_ff",
        "pipeline_a_ff",
        "pipeline_b_ff",
        "fitskin_L",
        "fitskin_a",
        "fitskin_b",
        "fairface_label",
        "predicted_ethnicity",
        "fairface_confidence",
        "n_cheek",
        "flash_scale",
        "zip_stem",
        "zip_path",
        "setting_label",
        "ff_sampling",
    ]
    fields = [c for c in preferred if c in fields] + [c for c in fields if c not in preferred]
    with csv_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    by_person = {}
    for r in rows:
        by_person.setdefault(r["person"], []).append(r["de00_off"])
    by_setting = {}
    for r in rows:
        by_setting.setdefault(r["setting"], []).append(r["de00_off"])

    summary = {
        "n_trials": len(rows),
        "n_failed": len(trials) - len(rows),
        "mean_de00_off": mean(r["de00_off"] for r in rows),
        "median_de00_off": median(r["de00_off"] for r in rows),
        "mean_de00_ff": mean(r["de00_ff"] for r in rows if "de00_ff" in r)
        if any("de00_ff" in r for r in rows)
        else None,
        "by_person_de00_off": {
            p: {"n": len(v), "mean": mean(v), "median": median(v)} for p, v in by_person.items()
        },
        "by_setting_de00_off": {
            s: {"n": len(v), "mean": mean(v), "median": median(v)} for s, v in sorted(by_setting.items())
        },
        "emily_note": "Emily has FitSkin Lab + planned settings in xlsx but no capture zips in this dataset.",
        "fixed_cat_k": args.fixed_cat_k,
        "cal_dir": str(args.cal_dir),
    }
    (args.out_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n")
    make_plots(rows, args.out_dir)

    print("\n=== Summary (frozen preawb_cat) ===")
    print(f"n={summary['n_trials']}  mean={summary['mean_de00_off']:.2f}  median={summary['median_de00_off']:.2f}")
    print(f"{'Person':10s} {'n':>4s} {'mean':>8s} {'median':>8s}")
    for p, st in summary["by_person_de00_off"].items():
        print(f"{p:10s} {st['n']:4d} {st['mean']:8.2f} {st['median']:8.2f}")
    print(f"{'Setting':10s} {'n':>4s} {'mean':>8s} {'median':>8s}")
    for s, st in summary["by_setting_de00_off"].items():
        print(f"{s:10s} {st['n']:4d} {st['mean']:8.2f} {st['median']:8.2f}")
    print("Wrote", csv_path)
    print("Plots in", args.out_dir)


if __name__ == "__main__":
    main()
